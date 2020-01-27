import openpyxl
from django.http import JsonResponse
from django.shortcuts import render


# Create your views here.


def admin_dashboard(request):
    if request.method == "GET":
        return render(request, 'dashboard/dashboard.html')
    else:
        excel_file = request.FILES['file']
        wb = openpyxl.load_workbook(excel_file)
        active_sheet = wb.active
        excel_data = []
        col_names = {}
        current = 0
        for col in active_sheet.iter_cols(1, active_sheet.max_column):
            col_names[col[0].value] = current
            current += 1
        for index, row in enumerate(active_sheet.iter_rows(min_row=2)):
            email = row[col_names['Email']].value
            description = row[col_names['Description']].value
            excel_data.append({"email": email, "description": description})
        return JsonResponse(excel_data, safe=False)
